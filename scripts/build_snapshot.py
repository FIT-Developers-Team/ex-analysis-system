"""Build the local gzip fast-path used by the Next.js data source.

This is intentionally read-only: it never changes the operational workbook.
"""

from __future__ import annotations

import argparse
import gzip
import json
import os
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


SHEET_TO_WH = {
    "Frozen - PGS": "PGS",
    "Frozen - SRG": "SRG",
    "Frozen - BIT": "BIT",
    "Frozen - STR": "STR",
}
SHEETS = [*SHEET_TO_WH, "Highlight"]


def iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 30000 < value < 70000:
        return (date(1899, 12, 30) + timedelta(days=float(value))).isoformat()
    if not isinstance(value, str):
        return None
    normalized = value.strip()
    if len(normalized) >= 10 and normalized[4:5] == "-" and normalized[7:8] == "-":
        return normalized[:10]
    try:
        return datetime.fromisoformat(normalized).date().isoformat()
    except ValueError:
        return None


def number_value(value: Any) -> tuple[float | int | None, str]:
    if isinstance(value, bool):
        return int(value), "valid"
    if isinstance(value, (int, float)):
        return value, "valid"
    if isinstance(value, str) and value.strip().startswith("#"):
        return None, "formula_error"
    if value in (None, ""):
        return None, "blank"
    try:
        text = str(value).replace(",", "").strip()
        parsed = float(text.replace("%", ""))
        return (parsed / 100 if "%" in text else parsed), "valid"
    except ValueError:
        return None, "blank"


def matrix_points(name: str, rows: list[tuple[Any, ...]], today: str) -> list[dict[str, Any]]:
    warehouse = SHEET_TO_WH[name]
    header_index = next(
        (index for index, row in enumerate(rows) if sum(1 for cell in row[7:] if iso_date(cell)) >= 3),
        -1,
    )
    if header_index < 0:
        return []
    header = rows[header_index]
    dated_columns = [(index, parsed) for index, cell in enumerate(header) if index >= 7 and (parsed := iso_date(cell))]
    points: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        metric = str(row[3] or "").strip() if len(row) > 3 else ""
        if not metric:
            continue
        division = str(row[0] or "Other").strip() or "Other"
        role = str(row[1] or "All").strip() or "All"
        remarks = str(row[2] or "").strip()
        detail = str(row[4] or "").strip()
        source = str((row[5] if len(row) > 5 else None) or (row[6] if len(row) > 6 else None) or name).strip()
        for column, metric_date in dated_columns:
            value, quality = number_value(row[column] if column < len(row) else None)
            if metric_date > today and quality == "valid":
                quality = "future"
            points.append({
                "warehouse": warehouse,
                "date": metric_date,
                "division": division,
                "role": role,
                "remarks": remarks,
                "metric": metric,
                "detail": detail,
                "source": source,
                "value": value,
                "quality": quality,
            })
    return points


def highlights(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    context_date: str | None = None
    records: list[dict[str, Any]] = []
    for row in rows:
        values = list(row) + [None] * max(0, 4 - len(row))
        maybe_date = iso_date(values[0])
        if maybe_date and not str(values[1] or "").strip():
            context_date = maybe_date
            continue
        warehouse = str(values[0] or "").strip()
        issue = str(values[2] or "").strip()
        if not warehouse or not issue or warehouse.lower() == "wh":
            continue
        records.append({
            "date": context_date,
            "warehouse": warehouse,
            "metric": str(values[1] or "").strip(),
            "issue": issue,
            "actionPlan": str(values[3] or "").strip(),
        })
    return records


def build(workbook_path: Path, output_path: Path) -> None:
    fetched_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    today = fetched_at[:10]
    workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    all_points: list[dict[str, Any]] = []
    highlight_rows: list[tuple[Any, ...]] = []
    try:
        for name in SHEETS:
            if name not in workbook.sheetnames:
                continue
            sheet = workbook[name]
            source_max_row = sheet.max_row
            source_max_col = sheet.max_column
            if source_max_row is None or source_max_col is None:
                _, _, source_max_col, source_max_row = range_boundaries(sheet.calculate_dimension(force=True))
            max_row = min(source_max_row, 500 if name == "Highlight" else 400)
            max_col = 4 if name == "Highlight" else min(max(source_max_col, 64), 500)
            rows = list(sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True))
            if name == "Highlight":
                highlight_rows = rows
            else:
                all_points.extend(matrix_points(name, rows, today))
    finally:
        workbook.close()

    valid_dates = sorted({point["date"] for point in all_points if point["quality"] == "valid"})
    payload = {
        "sourceMode": "workbook",
        "sourceName": workbook_path.name,
        "fetchedAt": fetched_at,
        "points": [point for point in all_points if point["quality"] in ("valid", "future")],
        "highlights": highlights(highlight_rows),
        "diagnostics": {
            "totalCells": len(all_points),
            "validCells": sum(point["quality"] == "valid" for point in all_points),
            "blankCells": sum(point["quality"] == "blank" for point in all_points),
            "formulaErrors": sum(point["quality"] == "formula_error" for point in all_points),
            "futureCells": sum(point["quality"] == "future" for point in all_points),
            "latestCompleteDate": valid_dates[-1] if valid_dates else None,
        },
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_suffix(f"{output_path.suffix}.tmp")
    with gzip.open(temporary_path, "wt", encoding="utf-8", compresslevel=6) as handle:
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
    temporary_path.replace(output_path)
    print(json.dumps({"snapshot": str(output_path), "points": len(all_points), "latest": payload["diagnostics"]["latestCompleteDate"]}))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?", type=Path, default=os.environ.get("FIT_WORKBOOK_PATH"))
    parser.add_argument("--output", type=Path, default=Path(".cache/operational-dataset.json.gz"))
    args = parser.parse_args()
    if not args.workbook:
        parser.error("pass a workbook path or set FIT_WORKBOOK_PATH")
    build(Path(args.workbook).resolve(), args.output.resolve())
